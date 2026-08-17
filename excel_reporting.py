"""Excel workbook output intended as the primary non-developer interface."""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from .ranking import RankedStock
from .config import CATEGORY_MAX, MODEL_VERSION

def _rank_rows(ranked: list[RankedStock]) -> list[dict]:
    rows=[]
    for r in ranked:
        rows.append({"Rank":r.rank,"Ticker":r.company.ticker,"Company":r.company.company_name,"Score":r.score.total_score,
                     "Classification":r.score.classification,"Price":r.company.share_price,"Risk Level":r.risks.severity,
                     "Data Confidence":r.quality.confidence,"Data Completeness":r.quality.completeness,
                     "Strengths":"; ".join(r.strengths),"Weaknesses":"; ".join(r.weaknesses),
                     "Risks":"; ".join(r.risks.flags),"Catalysts":"; ".join(r.company.catalysts)})
    return rows

def export_workbook(ranked: list[RankedStock], path="reports/investment_dashboard.xlsx", recommended: RankedStock | None=None) -> Path:
    p=Path(path); p.parent.mkdir(parents=True,exist_ok=True)
    current=pd.DataFrame(_rank_rows(ranked))
    scores=pd.DataFrame([{"Ticker":r.company.ticker, **{k:getattr(r.score,k) for k in CATEGORY_MAX}, "Total":r.score.total_score} for r in ranked])
    fundamentals=pd.DataFrame([{k:v for k,v in r.company.to_dict().items() if k not in {"metadata","catalysts","risk_factors"}} for r in ranked])
    quality=pd.DataFrame([{"Ticker":r.company.ticker,"Confidence":r.quality.confidence,"Completeness":r.quality.completeness,
                           "Level":r.quality.level,"Missing Fields":", ".join(r.quality.missing_fields),"Conflicts":", ".join(r.quality.conflicts),
                           "Warnings":r.quality.warning_count,"Errors":r.quality.error_count,"Sources":"; ".join(r.company.data_sources)} for r in ranked])
    decision=pd.DataFrame([{"Model Version":MODEL_VERSION,"Recommended Ticker":recommended.company.ticker if recommended else "NO ELIGIBLE CANDIDATE",
                            "Score":recommended.score.total_score if recommended else None,"Classification":recommended.score.classification if recommended else None,
                            "Risk":recommended.risks.severity if recommended else None,"Data Confidence":recommended.quality.confidence if recommended else None,
                            "Reason":"Highest-ranked candidate passing independent risk and data-confidence screens" if recommended else "All candidates blocked by risk/data-quality screen"}])
    with pd.ExcelWriter(p,engine="openpyxl") as writer:
        current.to_excel(writer,index=False,sheet_name="Current Rankings")
        scores.to_excel(writer,index=False,sheet_name="Category Scores")
        fundamentals.to_excel(writer,index=False,sheet_name="Fundamentals")
        quality.to_excel(writer,index=False,sheet_name="Data Quality")
        decision.to_excel(writer,index=False,sheet_name="Payday Decision")
    wb=load_workbook(p)
    for ws in wb.worksheets:
        ws.freeze_panes="A2"
        ws.auto_filter.ref=ws.dimensions
        for cell in ws[1]:
            cell.font=Font(bold=True); cell.fill=PatternFill("solid", fgColor="D9EAF7"); cell.alignment=Alignment(horizontal="center")
        for column in ws.columns:
            width=min(45,max(10,max(len(str(c.value or "")) for c in column)+2))
            ws.column_dimensions[column[0].column_letter].width=width
    if "Score" in [c.value for c in wb["Current Rankings"][1]]:
        col=[c.value for c in wb["Current Rankings"][1]].index("Score")+1
        letter=wb["Current Rankings"].cell(1,col).column_letter
        wb["Current Rankings"].conditional_formatting.add(f"{letter}2:{letter}{max(2,wb['Current Rankings'].max_row)}",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    wb.save(p); return p
