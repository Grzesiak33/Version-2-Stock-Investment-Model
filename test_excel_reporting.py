from openpyxl import load_workbook
from src.models import CompanyData
from src.ranking import rank_companies, recommended_candidate
from src.excel_reporting import export_workbook

def test_excel_dashboard_created(tmp_path):
    company=CompanyData("ABC",company_name="ABC Corp",share_price=10,revenue_growth=.2,free_cash_flow=10,cash=20,debt=1,
                        data_sources=["test"],data_timestamp="2026-08-17T12:00:00+00:00")
    ranked=rank_companies([company])
    p=export_workbook(ranked,tmp_path/"dashboard.xlsx",recommended_candidate(ranked,min_confidence=0))
    wb=load_workbook(p,read_only=True)
    assert {"Current Rankings","Category Scores","Fundamentals","Data Quality","Payday Decision"}.issubset(wb.sheetnames)
